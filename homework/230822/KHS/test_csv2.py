import csv
import glob
import os
import openpyxl

# 전역 변수부
file_list = glob.glob(os.path.join('c:/cookAnalysis/Excel/user/', '*.xlsx'))
firstYN = True
# 메인 코드부
for input_file in file_list:
    workbook = openpyxl.load_workbook(input_file)
    wsheetList = workbook.sheetnames
    for wsName in wsheetList:
        worksheet = workbook[wsName]
        with open("C:/CookAnalysis/Excel/유저통합_KHS.csv", "a", newline='') as outFp:
            csvWriter = csv.writer(outFp)
            if firstYN == True:
                header_list = []
                for col in range(1, worksheet.max_column + 1):
                    header_list.append(worksheet.cell(row=1, column=col).value)
                csvWriter.writerow(header_list)
                firstYN = False
            for row in range(2, worksheet.max_row + 1):
                row_list = []
                for col in range(1, worksheet.max_column + 1):
                    row_list.append(worksheet.cell(row=row, column=col).value)
                csvWriter.writerow(row_list)

print('Save. OK~')
